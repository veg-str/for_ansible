import json
import openpyxl
import re
import jinja2
import datetime
import zipfile
from pprint import pprint
from shared import ip_plan, sig_int

conf_template = 'project_files\\psm_config_template.json'
new_conf_path = 'c:\\temp\\psm_conf\\'

now = str(datetime.datetime.today().isoformat(sep="_", timespec="minutes")).replace(":", "_")

env = jinja2.Environment(
    loader=jinja2.FileSystemLoader('project_files')
)

#mr = ['SPB']
mr = ['EKT', 'NIN', 'NSK', 'MOS', 'SPB']

defaultDestinationRealm = 'bercut'
originRealm = 'node.epc.mnc020.mcc250.3gppnetwork.org'
max_counts = {
    'MOS': {'session_max_count': 1800000, 'subscriber_max_count': 16000000},
    'EKT': {'session_max_count': 1500000, 'subscriber_max_count': 6000000},
    'NIN': {'session_max_count': 1500000, 'subscriber_max_count': 6000000},
    'NSK': {'session_max_count': 1500000, 'subscriber_max_count': 6000000},
    'SPB': {'session_max_count': 1500000, 'subscriber_max_count': 6000000},
    }



def prepare_template(source):
    parts = source['components']
    remove_list = ['psm.diameterdmanager', 'psm.provisionerng.host']
    new_list = []
    for part in parts:
        if part['componentId'] not in remove_list:
            new_list.append(part)
    source['components'] = new_list
    return source


def get_cluster_list(srvs):
    cluster_list = []
    for srv in srvs:
        if re.search("^psm\d\d\..*\(VRRP VIP\)", srv['vm']) and srv['vlan'] == 'Radius':
            cluster_list.append(srv['vm'][:9])
    return cluster_list


def get_psm_vip(srv):
    wb = openpyxl.load_workbook(file_ip_plan, True)
    ws = wb[region]
    psm_vip = {}
    for row in ws.iter_rows():
        if re.search(srv + ".*\(VRRP VIP\)", str(row[1].value)):
            psm_vip[row[2].value] = row[5].value
    return psm_vip


def get_radius_secret(region):
    wb = openpyxl.load_workbook(file_sig_int, True)
    ws = wb[region]
    radius_secret = ws['M6'].value
    return radius_secret


def get_gy_peers(mr):
    wb = openpyxl.load_workbook(file_sig_int, True)
    s1_gy = wb.defined_names[mr.lower() + '_s1_gy'].attr_text
    ws = wb[mr]
    rng = s1_gy[s1_gy.find('!') + 1:]
    gy_peers = []
    for row in ws[rng]:
        gy_peers.append({"peerId": row[13].value,
                         "hostName": row[9].value,
                         "port": int(row[11].value),
                         "bindAddress": None,
                         "enabled": True,
                         "watchdogTimeoutMs": 30000
                         })
    gy_peers.append({"peerId": "T2TST-RTUCG-01-2",
                     "hostName": "10.78.245.57",
                     "port": 3878,
                     "bindAddress": None,
                     "enabled": False,  # True,
                     "watchdogTimeoutMs": 30000
                     })
    return gy_peers


def get_gx_config(srv, peer_list, region, conf_type):
    if int(re.search('\d\d', srv).group(0)) % 2 != 0:
        part = 'odd'
        env.globals['originRealm'] = region.lower() + '01.' + originRealm
    else:
        part = 'even'
        env.globals['originRealm'] = region.lower() + '02.' + originRealm
    template = env.get_template(f'{conf_type}_gx_config.txt.j2')
    env.globals['originHost'] = srv[:5]
    env.globals['gxVIP'] = psm_vip['Gx']
    env.globals['peer_list'] = peer_list[part]
    gx_config = template.render()
    return gx_config


def edit_psm_schema(config, srv, mr):
    obj = config['objects']
    for item in obj:
        if item['name'] == 'session':
            item['maxCount'] = max_counts[mr]['session_max_count']
            fields = obj[obj.index(item)]['fields']
            for field in fields:
                if field[1]['name'] == 'psmOriginHost':
                    field[1]['defaultValue'] = srv + '.' + originRealm
        if item['name'] == 'subscriber':
            item['maxCount'] = max_counts[mr]['subscriber_max_count']
    config['objects'] = obj
    messages = config['messages']
    for msg in messages:
        if msg['name'] == 'tele2cdr':
            fields = messages[messages.index(msg)]['fields']
            for field in fields:
                if field['name'] == 'nodeId':
                    field['defaultValue'] = srv
    config['messages'] = messages
    return config


def edit_psm_provisionerng_base(config, srv):
    fields = config['schemas'][0]['fields']
    for field in fields:
        if field['name'] == 'psmIdentity':
            field['mapping'] = '\"' + srv[:5] + '\"'
    config['schemas'][0]['fields'] = fields
    return config


def set_psm_diameterdmanager(gx_config):
    element = {'componentId': 'psm.diameterdmanager',
               'instanceId': None,
               'version': 5,
               'config': {'config': gx_config}
               }
    return element


def set_psm_provisionerng_host(pre):
    element = {'componentId': 'psm.provisionerng.host',
               'instanceId': None,
               'version': 4,
               'config': {'hostName': pre['ip'],
                          'serviceOrPort': '',
                          'useDirectConnection': False,
                          'startupState': 'enabled',
                          'group': ''}
               }
    return element


for region in mr:
    print(f'Generating config files for {region}.')
    sheet_list = ip_plan.get_sheets_list(region)
    zip_file = zipfile.ZipFile(f'{new_conf_path}psm_{region.upper()}_{now}.zip', 'w')
    for sheet in sheet_list:
        full_srv_list = ip_plan.rows_to_dict(sheet)
        psm_list = list(filter(ip_plan.check_psm, full_srv_list))
        pre_list = list(filter(ip_plan.check_pre, full_srv_list))
        pre_list = list(filter(lambda i: i['vlan'] == 'Provisioning', pre_list))
        provisioner = []
        for pre in pre_list:
            provisioner.append(set_psm_provisionerng_host(pre))
        gx_peers_local = sig_int.get_gx_peers(sheet, 'local_pcrf')
        gx_peers_dra = sig_int.get_gx_peers(sheet, 'dra')
        gy_peers = sig_int.get_gy_peers(sheet)
        radius_secret = sig_int.get_radius_secret(sheet)
        #        pprint(radius_secret)
        cl_list = get_cluster_list(psm_list)
        #        pprint(psm_list)
        with open(conf_template, "r", newline='\n') as file:
            config_template = json.load(file)
            for srv in cl_list:
                config_source = prepare_template(config_template)
                psm_vip = ip_plan.get_psm_vip(srv, psm_list)
                #                pprint(psm_vip)
                # **** Editing different component parameters ****
                components = config_source['components']
                for component in components:
                    if component['componentId'] == 'psm.diameter':
                        component['config']['originHost'] = srv[:9]
                        for gy_session in component['config']['gySessions']:
                            gy_session['defaultDestinationRealm'] = defaultDestinationRealm
                    elif component['componentId'] == 'psm.diameter.peers':
                        component['config']['routes'][0]['routes'][0]['destinationRealm'] = defaultDestinationRealm
                        peers = []
                        for peer in gy_peers:
                            peer["bindAddress"] = psm_vip['Gy']
                            peers.append(peer['peerId'])
                        component['config']['routes'][0]['routes'][0]['peerIds'] = peers
                        component['config']['peers'] = gy_peers
                    elif component['componentId'] == 'psm.model.syncer.client':
                        syncer_client = []
                        for psm in psm_list:
                            if srv in psm['vm'] and psm['vlan'] == 'ClusterSync':
                                syncer_client.append({'hostName': psm['ip'],
                                                      'serviceOrPort': "",
                                                      'useDirectConnection': False})
                        component['config']['nodes'] = syncer_client
                    elif component['componentId'] == 'psm.schema':
                        component['config'] = edit_psm_schema(component['config'], srv, region)
                    elif component['componentId'] == 'psm.source.udp.radius':
                        component['config']['hostName'] = psm_vip['Radius']
                        component['config']['secrets'][0]['secret'] = radius_secret
                    elif component['componentId'] == 'psm.provisionerng.base':
                        component['config']['identity'] = srv[:5]
                        component['config'] = edit_psm_provisionerng_base(component['config'], srv)
                # **** Adding remote and local Gx configuration sections ****
                components.append(set_psm_diameterdmanager(get_gx_config(srv, gx_peers_dra, region, 'remote')))
                components.append(set_psm_diameterdmanager(get_gx_config(srv, gx_peers_local, region, 'local')))
                # **** Adding Provisioning component for all PRE in region ****
                for pre in provisioner:
                    components.append(pre)
                # **** Writing new config to JSON file ****
                new_config_file = f'{new_conf_path}{srv.replace(".", "_")}_{now}.json'
                with open(new_config_file, 'w', newline='\n') as new_cfg:
                    json.dump(config_source, new_cfg)
                zip_file.write(new_config_file, f'{srv.replace(".", "_")}_{now}.json')
    zip_file.close()
