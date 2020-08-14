import openpyxl
import re

ip_plan = 'project_files\\Tele2_IP_plan_v2.04-draft.xlsx'
inventory_file = 'c:\\temp\\inventory\\kvm'

mr = ['MOS', 'NIN', 'EKT', 'NSK']
# mr = ['SPB', 'MOS', 'ROS', 'NIN', 'EKT', 'NSK']
base_srv_types = ['pre', 'psm', 'pic', 'apic']
ext_srv_types = ['epsm', 'rb', 'log', 'rs']

wb = openpyxl.load_workbook(ip_plan, True)


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def rows_to_dict(ws):
    rows = []
    ws = wb[ws]
    for row in ws.iter_rows():
        rows.append({
            'hostname': row[0].value,
            'vm': row[1].value,
            'vlan': row[3].value,
            'ip': row[5].value,
            'site': row[6].value,
            'domain': row[7].value
            })
    return rows


def chek_host_mgmt(item):
    if item['vlan'] == 'Host_Mgmt':
        return True


def get_vm_names(srv, srv_list):
    vm_name = []
    for row in srv_list:
        if row['hostname'] == srv['hostname'] and row['vlan'] == 'vm_Mgmt':
            vm_name.append(row['vm'][:-18])
    return vm_name


def site_groups(region, sheets, site):
    groups = []
    for type in base_srv_types + ext_srv_types:
        members = []
        i = 0
        while i < len(sheets):
            members.append(f'kvm_{region.lower()}{site[-1]}_d{i+1}_{type}')
            i = i + 1
        groups.append({'name': f'[kvm_{region.lower()}{site[-1]}_{type}:children]', 'members': members})
    return groups


def mr_groups(region, type, sites):
    members = []
    for site in sorted(sites):
        members.append(f'kvm_{region.lower()}{site[-1]}_{type}')
    group = {'name':f'[kvm_{region.lower()}_{type}:children]', 'members': members}
    return group


def domain_groups(region, site):
    members = []
    for type in base_srv_types:
        members.append(f'kvm_{region.lower()}{site[-1]}_d{dom_num}_{type}')
    groups = {'name': f'[kvm_{region.lower()}{site[-1]}_d{dom_num}:children]', 'members': members}
    return groups


def global_group(type):
    members = []
    for region in mr:
        members.append(f'kvm_{region.lower()}_{type}')
    group = {'name': f'[{type}:children]', 'members': members}
    return group


with open(inventory_file, 'w', newline='\n') as f:
    f.write('# List of KVMs\n\n')
    for region in mr:
        print(f'Collecting data about KVM in {region}')
        f.write(f'# {region}\n')
        sheet_list = get_sheets_list(region)
        mr_sites = set()
        for sheet in sheet_list:
            dom_num = sheet_list.index(sheet) + 1
            f.write(f'# Domain {dom_num}\n')
            kvm_list = rows_to_dict(sheet)
            hosts = list(filter(chek_host_mgmt, kvm_list))
            sites = sorted(set(map(lambda i: i['site'], hosts)))
            for site in sites:
                for type in base_srv_types:
                    f.write(f'[kvm_{region.lower()}{site[-1]}_d{dom_num}_{type}]\n')
                    for host in hosts:
                        vm_list = get_vm_names(host, kvm_list)
                        if host['site'] == site and re.search("^" + type, vm_list[0]):
                            f.write(f'{host["hostname"][:-13]} ansible_host={host["ip"]} '
                                    f'vm_name={vm_list[0]} vm_list={vm_list}\n')
                    f.write('\n')
                for type in ext_srv_types:
                    f.write(f'[kvm_{region.lower()}{site[-1]}_d{dom_num}_{type}]\n')
                    for host in hosts:
                        vm_list = get_vm_names(host, kvm_list)
                        if host['site'] == site and re.search(type + '\d{2}', str(vm_list)):
                            f.write(f'{host["hostname"][:-13]}\n')
                    f.write('\n')
                group = domain_groups(region, site)
                f.write(f'{group["name"]}\n')
                for member in group['members']:
                    f.write(f'{member}\n')
                f.write('\n')
                f.write(f'{group["name"][:-9]}vars]\n')
                f.write(f'dom_num={dom_num}\n\n')
                mr_sites.add(site)
        f.write(f'# Groups for {region}\n')
        for site in sorted(mr_sites):
            s_groups = site_groups(region, sheet_list, site)
            for group in s_groups:
                f.write(f'{group["name"]}\n')
                for member in group['members']:
                    f.write(f'{member}\n')
                f.write('\n')
        for srv_type in base_srv_types + ext_srv_types:
            group = mr_groups(region, srv_type, mr_sites)
            f.write(group['name'] + '\n')
            for member in group['members']:
                f.write(member + '\n')
            f.write('\n')
    f.write('\n# Global groups\n\n')
    for srv_type in base_srv_types + ext_srv_types:
        group = global_group(srv_type)
        f.write(group['name'] + '\n')
        for member in group['members']:
            f.write(member + '\n')
        f.write('\n')
    wb.close()
    print('Done')