import openpyxl
import re

ip_plan = 'project_files/Tele2_IP_plan_v3.01.xlsx'
inventory_file = 'c:/temp/inventory/packetlogick'

mr = ['MOS', 'NIN', 'EKT', 'NSK']
#mr = ['SPB', 'MOS', 'ROS', 'NIN', 'EKT', 'NSK']
base_srv_types = ['pre', 'psm', 'pic', 'apic']
ext_srv_types = ['epsm', 'rb', 'log', 'rs']

wb = openpyxl.load_workbook(ip_plan, True)


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def domain_groups(region, site):
    members = []
    for type in base_srv_types:
        members.append(f'pl_{region.lower()}{site[-1]}_d{dom_num}_{type}')
    groups = {'name': f'[pl_{region.lower()}{site[-1]}_d{dom_num}:children]', 'members': members}
    return groups


def site_groups(region, sheets, site):
    groups = []
    for type in base_srv_types:
        members = []
        i = 0
        while i < len(sheets):
            members.append(f'pl_{region.lower()}{site[-1]}_d{i+1}_{type}')
            i = i + 1
        groups.append({'name': f'[pl_{region.lower()}{site[-1]}_{type}:children]', 'members': members})
    return groups


def mr_groups(region, type, sites):
    members = []
    for site in sorted(sites):
        members.append(f'pl_{region.lower()}{site[-1]}_{type}')
    group = {'name':f'[pl_{region.lower()}_{type}:children]', 'members': members}
    return group


def global_group(type):
    members = []
    for region in mr:
        members.append(f'pl_{region.lower()}_{type}')
    group = {'name': f'[{type}:children]', 'members': members}
    return group


def pre_prov_ip(srv, srv_list):
    prov_ip = ''
    for row in srv_list:
        if row['hostname'] == srv['hostname'] and row['vlan'] == 'Provisioning':
            prov_ip = row['ip']
    return prov_ip


def pic_dadafeed_ip(srv, srv_list):
    df_ip = ''
    for row in srv_list:
        if row['hostname'] == srv['hostname'] and row['vlan'] == 'DataFeed':
            df_ip = row['ip']
    return df_ip


def rows_to_dict(sheet):
    rows = []
    ws = wb[sheet]
    for row in ws.iter_rows():
        if re.search("^\D{1,4}", str(row[1].value)).group(0) in base_srv_types:
            rows.append({
                'hostname': row[1].value,
                'vlan': row[3].value,
                'ip': row[5].value,
                'site': row[6].value})
    return rows


with open(inventory_file, 'w', newline='\n') as f:
    f.write('# List of PacketLogic VMs\n\n')
    for region in mr:
        print(f'Collecting data about VM in {region}')
        sheet_list = get_sheets_list(region)
        mr_sites = set()
        for sheet in sheet_list:
            dom_num = sheet_list.index(sheet) + 1
            f.write(f'# Domain {dom_num}\n')
            vm_list = rows_to_dict(sheet)
            sites = sorted(set(map(lambda i: i['site'], vm_list)))
            for site in sites:
                for type in base_srv_types:
                    f.write(f'[pl_{region.lower()}{site[-1]}_d{dom_num}_{type}]\n')
                    for vm in vm_list:
                        if (vm['vlan'] == 'vm_Mgmt' and
                                vm['site'] == site and
                                re.search("^" + type, vm['hostname'])):
                            f.write(f'{vm["hostname"][:-13]} ansible_host={vm["ip"]}')
                            if type == 'pre':
                                f.write(f' provisioning_ip={pre_prov_ip(vm, vm_list)}\n')
                            elif type in ['pic', 'apic']:
                                f.write(f' datafeed_ip={pic_dadafeed_ip(vm, vm_list)}\n')
                            else:
                                f.write('\n')
                    f.write('\n')
                group = domain_groups(region, site)
                f.write(f'{group["name"]}\n')
                for member in group['members']:
                    f.write(f'{member}\n')
                f.write('\n')
                f.write(f'{group["name"][:-9]}vars]\n')
                f.write(f'dom_num={dom_num}\n\n')
                mr_sites.add(site)
        f.write(f'# Groups for {region}\n')
        for site in sorted(mr_sites):
            s_groups = site_groups(region, sheet_list, site)
            for group in s_groups:
                f.write(f'{group["name"]}\n')
                for member in group['members']:
                    f.write(f'{member}\n')
                f.write('\n')
        for srv_type in base_srv_types:
            group = mr_groups(region, srv_type, mr_sites)
            f.write(group['name'] + '\n')
            for member in group['members']:
                f.write(member + '\n')
            f.write('\n')
    f.write('\n# Global groups\n\n')
    for srv_type in base_srv_types:
        group = global_group(srv_type)
        f.write(group['name'] + '\n')
        for member in group['members']:
            f.write(member + '\n')
        f.write('\n')
    wb.close()
    print('Done')
