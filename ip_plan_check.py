import openpyxl
import ipaddress
import re

ip_plan = 'project_files\\Tele2_IP_plan_v3.04_draft.xlsx'
#mr = ['mos']
mr = ['spb', 'mos', 'nin', 'ekt', 'nsk', 'ros']
check_results = {}

wb = openpyxl.load_workbook(ip_plan)
error_cell = openpyxl.styles.Font(color="FF0000")
has_errors = False


def get_def_names(region):
    wsid = wb.sheetnames.index('IP-plan')
    dn_list = list(filter(lambda i: re.search('^'+region, i), wb.defined_names.localnames(wsid)))
    return dn_list


def get_vlans():
    vlans = []
    ws = wb['Dictionary']
    i = 2
    cell = 'D' + str(i)
    while ws[cell].value:
        if 'FlowControl' not in ws[cell].value:
            vlans.append(ws[cell].value)
        i = i + 1
        cell = 'D' + str(i)
    print(vlans)
    return vlans


def check_subnets(region, dn_list):
    global has_errors
    test_result = '\033[32mPASSED\033[30m'
    print(f'Checking subnets in region {region.upper()}... ', end =" ")
    for dn in dn_list:
        wsid = wb.sheetnames.index('IP-plan')
        net = wb.defined_names.get(dn, wsid).attr_text
        ws = wb[str(net[1:net.find('!') - 1])]
        rng = net[net.find('!') + 1:]
        for i in 5, 8:
            for row in ws[rng]:
                try:
                    if row[0].value == 'Supernet':
                        supernet = ipaddress.ip_network(row[i].value + row[3].value)
                    elif row[0].value in vlans or row[0].value == 'Link subnets':
                        net = ipaddress.ip_network(row[i].value + row[3].value)
                        if not net.subnet_of(supernet):
                            row[i].font = error_cell
                            test_result = '\033[31mFAILED\033[30m'
                            has_errors = True
                except(ValueError, TypeError):
                    row[i].font = error_cell
                    test_result = '\033[31mFAILED\033[30m'
                    has_errors = True
    print(test_result)
    return test_result


def check_gw(region, dn_list):
    global has_errors
    test_result = '\033[32mPASSED\033[30m'
    print(f'Checking gateways in region {region.upper()}... ', end =" ")
    for dn in dn_list:
        wsid = wb.sheetnames.index('IP-plan')
        net = wb.defined_names.get(dn, wsid).attr_text
        ws = wb[str(net[1:net.find('!') - 1])]
        rng = net[net.find('!') + 1:]
        for row in ws[rng]:
            try:
                if row[0].value in vlans:
                    gw_index = -2
                    for i in 5, 8:
                        if i == 8 and row[5].value == row[8].value:
                            gw_index = -3
                        net = ipaddress.ip_network(row[i].value + row[3].value)
                        gw = ipaddress.ip_address(row[i+1].value)
                        gw_chek = net[gw_index]
                        if ipaddress.ip_address(row[i+1].value) != net[gw_index]:
                            row[i+1].font = error_cell
                            test_result = '\033[31mFAILED\033[30m'
                            has_errors = True
            except(ValueError, TypeError):
                row[i].font = error_cell
                test_result = '\033[31mFAILED\033[30m'
                has_errors = True
    print(test_result)
    return test_result


def check_ip(region, vlan, dn_list):
    global has_errors
    test_result = '\033[32mPASSED\033[30m'
    print(f'Checking IPs in VLAN {vlan} in region {region.upper()}... ', end =" ")
    for dn in dn_list:
        dom = re.search("[0-9]+$", str(dn)).group(0)
        wsid = wb.sheetnames.index('IP-plan')
        net = wb.defined_names.get(dn, wsid).attr_text
        ws = wb[str(net[1:net.find('!') - 1])]
        rng = net[net.find('!') + 1:]
        for row in ws[rng]:
            if row[0].value == vlan:
                net1 = ipaddress.ip_network(row[5].value + row[3].value)
                gw1 = ipaddress.ip_address(row[5].value)
                net2 = ipaddress.ip_network(row[8].value + row[3].value)
                gw2 = ipaddress.ip_address(row[9].value)
        ws = wb[f'{region.upper()}_D{dom}']
        for row in ws.iter_rows():
            if row[3].value == vlan:
                if row[6].value == 'Site1':
                    net = net1
                    gw = gw1
                elif row[6].value == 'Site2':
                    net = net2
                    gw = gw2
                if ipaddress.ip_address(row[5].value) not in net.hosts() or ipaddress.ip_address(row[5].value) == gw:
                    row[5].font = error_cell
                    test_result = '\033[31mFAILED\033[30m'
                    has_errors = True
    print(test_result)
    return test_result


def check_uniq_ip(region, vlan, dn_list):
    global has_errors
    test_result = '\033[32mPASSED\033[30m'
    print(f'Checking for IPs uniq in VLAN {vlan} in region {region.upper()}... ', end =" ")
    for dn in dn_list:
        dom = re.search("[0-9]+$", str(dn)).group(0)
        ws = wb[f'{region.upper()}_D{dom}']
        addr = []
        for row in ws.iter_rows():
            if row[3].value == vlan:
                addr.append(row[5].value)
        addr_set = set(addr)
        if len(addr) != len(addr_set):
            test_result = '\033[31mFAILED\033[30m'
            has_errors = True
    print(test_result)
    return test_result


vlans = get_vlans()
for region in mr:
    dn = get_def_names(region)
    check_results['Subnets'] = check_subnets(region, dn)
    check_results['GW'] = check_gw(region, dn)
    if not has_errors:
        for vlan in vlans:
            check_results['IP_' + vlan] = check_ip(region, vlan, dn)
        for vlan in vlans:
            check_results['Uniq_IP_' + vlan] = check_uniq_ip(region, vlan, dn)
#    print(f'Check results for {region.upper()}:' )
#    for test in check_results:
#        print(f'{test}: {check_results[test]}')

if has_errors:
    print('\033[31mATTENTION!!! File has errors!\033[30m')

wb.save(ip_plan)
wb.close()
