import openpyxl
import yaml
import re

ip_plan = 'project_files\\Tele2_IP_plan_v2.04-draft.xlsx'
vars_file = 'c:\\temp\\group_vars\\networks.yml'

wb = openpyxl.load_workbook(ip_plan, True)
ws_id = wb.sheetnames.index('IP-plan')

mr = ['MOS']
#mr = ['SPB', 'MOS', 'ROS', 'NIN', 'EKT', 'NSK']
reg = {}
subnets = {}


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def get_vlans():
    vlans = []
    ws = wb['Dictionary']
    i = 2
    cell = 'D' + str(i)
    while ws[cell].value:
        if 'FlowControl' not in ws[cell].value:
            vlans.append(ws[cell].value)
        i = i + 1
        cell = 'D' + str(i)
    print(vlans)
    return vlans


def get_sites_qnt():
    s_qnt = {}
    def_name = wb.defined_names['sites_qnt'].attr_text
    ws = wb[str(def_name[0:def_name.find('!')])]
    rng = def_name[def_name.find('!') + 1:]
    for row in ws[rng]:
        s_qnt[row[0].value] = int(row[1].value)
    return s_qnt


with open(vars_file, 'w', newline='\n') as f:
    vlans = get_vlans()
    sites_qnt = get_sites_qnt()
    f.write('#\n# Networks, prefixes and gateways in Tele2 TMS\n#\n')
    for region in mr:
        print(f'Collecting data about networks in {region}')
        sheet_list = get_sheets_list(region)
        for sheet in sheet_list:
            dom_num = sheet_list.index(sheet) + 1
            sites = [x for x in range(1, sites_qnt[sheet]+1)]
            net = wb.defined_names.get(sheet.lower(), ws_id).attr_text
            ws = wb[str(net[1:net.find('!') - 1])]
            rng = net[net.find('!') + 1:]
            for i in sites:
                site = {}
                if i == 1:
                    for row in ws[rng]:
                        if row[0].value in vlans:
                            site[row[0].value] = {'subnet': row[5].value, 'prefix': row[3].value, 'gw': row[6].value}
                else:
                    for row in ws[rng]:
                        if row[0].value in vlans:
                            site[row[0].value] = {'subnet': row[8].value, 'prefix': row[3].value, 'gw': row[9].value}
                reg[f'{region.lower()}{str(i)}_d{dom_num}'] = site
    subnets['network'] = reg
    f.write(yaml.dump(subnets))
    print('Done')
