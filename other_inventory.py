import openpyxl
import re

ip_plan = 'project_files\\Tele2_IP_plan_v2.04-draft.xlsx'
inventory_file = 'c:\\temp\\inventory\\other'

mr = ['MOS', 'NIN', 'EKT', 'NSK']
#mr = ['SPB', 'MOS', 'ROS', 'NIN', 'EKT', 'NSK']
base_srv_types = ['pre', 'psm', 'pic', 'apic']
ext_srv_types = ['epsm', 'rb', 'log', 'rs']

wb = openpyxl.load_workbook(ip_plan, True)


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def rows_to_dict(sheet):
    rows = []
    ws = wb[sheet]
    for row in ws.iter_rows():
        if re.search("^\D{1,4}", str(row[1].value)).group(0) in ext_srv_types:
            rows.append({
                'hostname': row[1].value,
                'vlan': row[3].value,
                'ip': row[5].value,
                'site': row[6].value
                })
    return rows


def get_prov_ip(srv, srv_list):
    prov_ip = ''
    for row in srv_list:
        if row['hostname'] == srv['hostname'] and row['vlan'] == 'Provisioning':
            prov_ip = row['ip']
    return prov_ip


def get_rad_ip(srv, srv_list):
    rad_ip = ''
    for row in srv_list:
        if row['hostname'] == srv['hostname'] and row['vlan'] == 'Radius':
            rad_ip = row['ip']
    return rad_ip


def domain_groups(region, site):
    members = []
    for type in ext_srv_types:
        members.append(f'oth_{region.lower()}{site[-1]}_d{dom_num}_{type}')
    groups = {'name': f'[oth_{region.lower()}{site[-1]}_d{dom_num}:children]', 'members': members}
    return groups


def site_groups(region, sheets, site):
    groups = []
    for type in ext_srv_types:
        members = []
        i = 0
        while i < len(sheets):
            members.append(f'oth_{region.lower()}{site[-1]}_d{i+1}_{type}')
            i = i + 1
        groups.append({'name': f'[oth_{region.lower()}{site[-1]}_{type}:children]', 'members': members})
    return groups


def mr_groups(region, type, sites):
    members = []
    for site in sorted(sites):
        members.append(f'oth_{region.lower()}{site[-1]}_{type}')
    group = {'name':f'[oth_{region.lower()}_{type}:children]', 'members': members}
    return group


def global_group(type):
    members = []
    for region in mr:
        members.append(f'oth_{region.lower()}_{type}')
    group = {'name': f'[{type}:children]', 'members': members}
    return group


with open(inventory_file, 'w', newline='\n') as f:
    f.write('# List of VMs\n\n')
    for region in mr:
        print(f'Collecting data about VM in {region}')
        f.write(f'# {region}\n')
        sheet_list = get_sheets_list(region)
        mr_sites = set()
        for sheet in sheet_list:
            dom_num = sheet_list.index(sheet) + 1
            f.write(f'# Domain {dom_num}\n')
            srv_list = rows_to_dict(sheet)
            sites = sorted(set(map(lambda i: i['site'], srv_list)))
            for site in sites:
                for type in ext_srv_types:
                    f.write(f'[oth_{region.lower()}{site[-1]}_d{dom_num}_{type}]\n')
                    for row in srv_list:
                        if (row['vlan'] == 'vm_Mgmt' and
                                row['site'] == site and
                                re.search("^" + type, row['hostname'])):
                            f.write(f'{row["hostname"][:-13]} ansible_host={row["ip"]}')
                            if type == 'epsm':
                                f.write(f' provisioning_ip={get_prov_ip(row, srv_list)}')
                                f.write(f' radius_ip={get_rad_ip(row, srv_list)}\n')
                            else:
                                f.write('\n')
                    f.write('\n')
                group = domain_groups(region, site)
                f.write(f'{group["name"]}\n')
                for member in group['members']:
                    f.write(f'{member}\n')
                f.write('\n')
                f.write(f'{group["name"][:-9]}vars]\n')
                f.write(f'dom_num={dom_num}\n\n')
                mr_sites.add(site)
        f.write(f'# Groups for {region}\n')
        for site in sorted(mr_sites):
            s_groups = site_groups(region, sheet_list, site)
            for group in s_groups:
                f.write(f'{group["name"]}\n')
                for member in group['members']:
                    f.write(f'{member}\n')
                f.write('\n')
        for srv_type in ext_srv_types:
            group = mr_groups(region, srv_type, mr_sites)
            f.write(group['name'] + '\n')
            for member in group['members']:
                f.write(member + '\n')
            f.write('\n')
    f.write('\n# Global groups\n\n')
    for srv_type in ext_srv_types:
        group = global_group(srv_type)
        f.write(group['name'] + '\n')
        for member in group['members']:
            f.write(member + '\n')
        f.write('\n')
    wb.close()
    print('Done')