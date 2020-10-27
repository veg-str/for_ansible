import openpyxl
import yaml
import ipaddress
import copy
import os
import re
from shared import sig_int

sig_int_file = 'project_files/Tele2_TMS_Signal_integration_v5.5.xlsx'
vars_dir = 'c:/temp/group_vars/'

# Open Excel file in read-only mode
wb = openpyxl.load_workbook(sig_int_file, True)

mr = ['MOS', 'NIN', 'NSK', 'EKT']
#mr = ['spb', 'nin', 'ekt', 'nsk', 'ros', 'mos']
#mr = ['nin', 'ekt', 'nsk', 'mos']
sigs = ['radius', 'gx', 'gy']

def make_var_dirs(dir_name):
    if dir_name not in os.listdir(vars_dir):
        os.mkdir(vars_dir + dir_name)
    dir = f'{vars_dir}{dir_name}'
    return dir


def filter_net_list(wsheet, sig, nets):
    fltr = f'{wsheet.lower()}_s\d_{sig}'
    filtered_nets = { key:value for (key, value) in nets.items() if re.match(fltr, key) }
    return filtered_nets


def summarise_nets(range): # Summarise addresses into /24 subnets
    supernets = set()
    for ip in range:
        supernets.add(str(ipaddress.IPv4Network(ip).supernet(new_prefix=24)))
    supernets = list(set(supernets))
    return supernets


def check_site_of_net(net, region, sig):
    for addr in signal_networks[region][sig]:
        if ipaddress.IPv4Network(addr['ip']).subnet_of(ipaddress.IPv4Network(net)):
            site = addr['site']
    return site


def get_psm_static_route(sig, nets):  # Signal Integration dependent part of routes
    s_routes = []
    next_hops = {
        'radius': {
            1: '{{ rb01_vip }}',
            2: '{{ rb02_vip }}'
        },
        'gx': '{{ networks[net_scope]["Gx1" if inventory_hostname_short[-2:] | int is odd else "Gx2"].gw }}',
        'gy': '{{ networks[net_scope]["Gy1" if inventory_hostname_short[-2:] | int is odd else "Gy2"].gw }}'
    }
    interfaces = {
        'radius': "Radius",
        'gx': "Gx",
        'gy': "Gy"
    }
    for key, value in nets.items():
        if sig == 'radius':
            site = int(re.search('_s\d', key).group(0)[-1])
            for net in value:
                sr = {
                    'route': net,
                    'next_hop': next_hops['radius'][site],
                    'interface': interfaces['radius']
                }
                if s_routes.count(sr) == 0:
                    s_routes.append(sr)
        else:
            for net in value:
                sr = {
                    'route': net,
                    'next_hop': next_hops[sig],
                    'interface': interfaces[sig]
                }
                if s_routes.count(sr) == 0:
                    s_routes.append(sr)
    return s_routes


def get_psm_const_sr():
    # Signal Integration independent part. Constant for all regions
    static_routes = []
    sr = {
        'route': "{{ networks[other_site_net_scope].Provisioning.subnet + networks[other_site_net_scope].Provisioning.prefix }}",
        'next_hop': "{{ networks[net_scope].Provisioning.gw }}",
        'interface': "Provisioning"
    }
    static_routes.append(sr)
    sr = {
        'route': "{{ networks[other_site_net_scope].ClusterSync.subnet + networks[other_site_net_scope].ClusterSync.prefix }}",
        'next_hop': "{{ networks[net_scope].ClusterSync.gw }}",
        'interface': "ClusterSync"
    }
    static_routes.append(sr)
    return static_routes


def get_rb_static_routes(nets):
    s_routes = []
    for key, value in nets.items():
        for net in value:
            sr = {
                'route': net,
                'next_hop': "{{ networks[net_scope].RadiusFE.gw }}",
                'interface': "enp3s0"
            }
            if s_routes.count(sr) == 0:
                s_routes.append(sr)
    return s_routes


def get_epsm_static_routes(nets):
    s_routes = []
    for key, value in nets.items():
        for net in value:
            sr = {
                'route': net,
                'next_hop': "{{ networks[net_scope].Radius.gw }}",
                'interface': "enp2s0"
            }
            if s_routes.count(sr) == 0:
                s_routes.append(sr)
    sr = {
        'route': "{{ networks[other_site_net_scope].Provisioning.subnet + networks[other_site_net_scope].Provisioning.prefix }}",
        'next_hop': "{{ networks[net_scope].Provisioning.gw }}",
        'interface': "enp3s0"
    }
    s_routes.append(sr)
    return s_routes

# Getting Signal Interation subnets
signal_networks = sig_int.get_sig_nets()

#  Get summarized subnets
sum_networks = copy.deepcopy(signal_networks)
keys = list(sum_networks.keys())
for key in keys:
        nets = summarise_nets(sum_networks[key])
        nets.sort()
        sum_networks[key] = nets

# Creates PSM group vars files with static routes
for region in mr:
    wsheets = sig_int.get_sheets_list(region)
    for ws in wsheets:
        psm_sr = []
        for sig in sigs:
            nets = filter_net_list(ws, sig, sum_networks)
            psm_sr.extend(get_psm_static_route(sig, nets))
            psm_sr.extend(get_psm_const_sr())
        dir = make_var_dirs(f'pl_{ws.lower()}_psm')
        print(f'Generating file for for PSMs in {ws}...', end='')
        with open(f'{dir}/static_routes.yml', 'w', newline='\n') as file:
            yaml.dump({'static_routes': psm_sr}, file, width=120, sort_keys=False)
        print(' Done')
print('')

# Create RB group vars files with static routes
for region in mr:
    wsheets = list(filter(lambda i: re.match(region, i, re.I), wb.sheetnames))
    for ws in wsheets:
        sig = 'radius'
        nets = filter_net_list(ws, sig, sum_networks)
        rb_sr = get_rb_static_routes(nets)
        dir = make_var_dirs(f'oth_{ws.lower()}_rb')
        print(f'Generating file for for RBs in {ws}...', end='')
        with open(f'{dir}/static_routes.yml', 'w', newline='\n') as file:
            yaml.dump({'static_routes': rb_sr}, file, width=120, sort_keys=False)
        print(' Done')
print('')

# Create ePSM group vars files with static routes
for region in mr:
    wsheets = list(filter(lambda i: re.match(region, i, re.I), wb.sheetnames))
    for ws in wsheets:
        sig = 'radius'
        nets = filter_net_list(ws, sig, sum_networks)
        rb_sr = get_epsm_static_routes(nets)
        dir = make_var_dirs(f'oth_{ws.lower()}_epsm')
        print(f'Generating file for for EPSMs in {ws}...', end='')
        with open(f'{dir}/static_routes.yml', 'w', newline='\n') as file:
            yaml.dump({'static_routes': rb_sr}, file, width=120, sort_keys=False)
        print(' Done')
print('')
