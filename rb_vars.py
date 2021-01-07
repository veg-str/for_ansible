import openpyxl
import yaml
import re

ip_plan = 'project_files/Tele2_IP_plan_v3.04_draft.xlsx'
vars_dir = 'c:/temp/host_vars/'

# Open Excel file in read-only mode
wb = openpyxl.load_workbook(ip_plan, True)

#mr = ['SPB']
#mr = ['MOS', 'NIN', 'EKT', 'NSK']
mr = ['SPB', 'MOS', 'ROS', 'NIN', 'EKT', 'NSK']


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def get_rb_list(sheet):
    ws = wb[sheet.upper()]
    rb_rows = []
    for row in ws.iter_rows():
        if re.search("^rb[0-9]+", str(row[1].value)):
            rb_rows.append({
                'vm_name': row[1].value,
                'vlan': row[3].value,
                'ip': row[5].value,
                'site': row[6].value
                })
    return rb_rows


def get_rb_vars(rb, rb_list):
    vars = {}
    vars['net'] = {}
    vars['ha'] = {}
    vars['ha']['vip'] = {}
    for srv in rb_list:
        if srv['vm_name'] == rb['vm_name']:
            if srv['vlan'] == 'Radius':
                vars['net']['radius_ip'] = srv['ip']
            elif srv['vlan'] == 'RadiusFE':
                vars['net']['radiusfe_ip'] = srv['ip']
        if re.search(' \(VRRP VIP\)', srv['vm_name']) and srv['site'] == rb['site']:
            if srv['vlan'] == 'Radius':
                vars['ha']['vip']['radius_vip'] = srv['ip']
            elif srv['vlan'] == 'RadiusFE':
                vars['ha']['vip']['radiusfe_vip'] = srv['ip']
    return vars


for region in mr:
    print(f'Collecting data about RB in {region}')
    sheet_list = get_sheets_list(region)
    for sheet in sheet_list:
        rb_list = get_rb_list(sheet)
        for rb in rb_list:
            if rb['vlan'] == 'vm_Mgmt':
                rb_vars = get_rb_vars(rb, rb_list)
                var_file = f'{vars_dir}{rb["vm_name"][:9]}.yml'
                with open(var_file, 'w', newline='\n') as f:
                    f.write(f'# Variables for {rb["vm_name"]}\n#\n')
                    f.write('# High availability related vars\n#\n')
                    f.write(yaml.dump(rb_vars['ha']))
                with open(var_file, 'a', newline='\n') as f:
                    f.write('#\n' + '# NICs related vars\n#\n')
                    f.write(yaml.dump(rb_vars['net']))
wb.close()
print('Done')