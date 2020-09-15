import openpyxl
import re

file_sig_int = 'project_files/Tele2_TMS_Signal_integration_v5.2.xlsx'
wb = openpyxl.load_workbook(file_sig_int, True)


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region.upper(), i), wb.sheetnames))
    return ws_list


def get_named_ranges(work_sheet):
    nr = wb.get_named_ranges()
    named_ranges = list(filter(lambda i: re.search('^'+work_sheet.lower(), i.name), nr))
    return named_ranges


def get_gy_peers(w_sheet):
    ws = wb[w_sheet]
    s1_gy = wb.defined_names[f'{w_sheet.lower()}_s1_gy'].attr_text
    rng = s1_gy[s1_gy.find('!') + 1:]
    gy_peers = []
    for row in ws[rng]:
        gy_peers.append({"peerId": row[13].value,
                         "hostName": row[9].value,
                         "port": int(row[11].value),
                         "bindAddress": None,
                         "enabled": True,
                         "watchdogTimeoutMs": 30000
                         })
    # Non-production RTUCG for testing purposes. Disabled by default
    gy_peers.append({"peerId": "T2TST-RTUCG-01-2",
                     "hostName": "10.78.245.57",
                     "port": 3878,
                     "bindAddress": None,
                     "enabled": False,  # True,
                     "watchdogTimeoutMs": 30000
                     })
    return gy_peers


def get_radius_secret(w_sheet):
    ws = wb[w_sheet]
    rad_secret = ws['M6'].value
    return rad_secret


def get_gx_peers(w_sheet, p_type):
    ws = wb[w_sheet]
    name = [None, 'odd', 'even']
    pcrfs = {}
    for i in 1,2:
        gx = wb.defined_names[f'{w_sheet.lower()}_s{i}_gx'].attr_text
        rng = gx[gx.find('!') + 1:]
        pcrfs_list = []
        for row in ws[rng]:
            if row[-1].value == p_type:
                pcrfs_list.append({
                    'primIP': row[9].value,
                    'secIP': row[10].value,
                    'hostName': row[13].value,
                    'realm': row[14].value
                })
        pcrfs[name[i]] = pcrfs_list
    return pcrfs
