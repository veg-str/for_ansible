import openpyxl
import re
from shared.classes import PlServer, PsmCluster

file_ip_plan = 'project_files/Tele2_IP_plan_v3.04_draft.xlsx'
base_srv_types = ['pre', 'psm', 'pic', 'apic']
ext_srv_types = ['epsm', 'rb', 'log', 'rs']

wb = openpyxl.load_workbook(file_ip_plan, True)


# New function with PlServer class usage

def rows_to_list(ws):
    full_srv_list = []
    for row in ws.iter_rows():
        full_srv_list.append({
            'hostname': str(row[0].value),
            'vm': str(row[1].value),
            'vlan': str(re.search('^[^0-9]*', str(row[3].value)).group(0)),
            'ip': str(row[5].value),
            'site': str(row[6].value)
        })
    return full_srv_list


def get_pl_list(region, domain=0):
    pl_list = []
    if domain == 0:
        ws_list = get_sheets_list(region)
    else:
        ws_list = [region.upper()+'_D'+str(domain)]
    for sheet in ws_list:
        ws = wb[sheet]
        srv_list = rows_to_list(ws)
        dom_num = int(sheet[-1])
        for srv in srv_list:
            if re.search('^[^0-9]{3,4}', srv['vm']):
                pl_type = re.search('^[^0-9]{3,4}', srv['vm']).group(0)
            else:
                pl_type = ''
            if srv['vlan'] == 'vm_Mgmt' and pl_type in base_srv_types:
                srv_ips = list(filter(lambda i: (i['vm'] == srv['vm'] and i['vlan'] != 'vm_Mgmt'), srv_list))
                pl_ips = {}
                for i in srv_ips:
                    pl_ips[i['vlan']] = i['ip']
                pl = PlServer(pl_type, srv['vm'], srv['ip'], region, srv['site'], dom_num, pl_ips)
                pl_list.append(pl)
    return pl_list


def get_psm_clusters(region, domain):
    clusters = []
    psm_list = filter_by_type(get_pl_list(region, domain), 'psm')
    psms = []
    for psm in psm_list:
        psms.append(psm.hostname[:-14])
    psms = sorted(set(psms))
    ws = wb[f'{region.upper()}_D{str(domain)}']
    for psm in psms:
        cl_name = psm
        cl_id = int(re.search('[0-9]{2}', psm).group(0))
        members = [f'{psm}1', f'{psm}2']
        vip = {}
        for row in ws.iter_rows():
            if re.search(psm + ' \(VRRP VIP\)', str(row[1].value)):
                vip[row[2].value] = row[5].value
        clusters.append(PsmCluster(cl_name, cl_id, members, vip))
    return clusters


def filter_by_type(srv_list, pl_type):
    filtered_list = list(filter(lambda i: i.pl_type == pl_type, srv_list))
    return filtered_list


def filter_by_site(srv_list, site):
    filtered_list = list(filter(lambda i: i.site == site, srv_list))
    return filtered_list


def filter_by_domain(srv_list, domain):
    filtered_list = list(filter(lambda i: i.domain == domain, srv_list))
    return filtered_list


# Old functions w/o classes

def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def rows_to_dict(sheet):
    rows = []
    ws = wb[sheet]
    for row in ws.iter_rows():
        rows.append({
            'hostname': row[0].value,
            'vm': row[1].value,
            'vlan': row[3].value,
            'ip': row[5].value,
            'site': row[6].value
        })
    return rows


def filter_host_only(src):
    res = src['vlan'] == 'Host_Mgmt'
    return res


def filter_vm_only(src):
    res = src['vlan'] == 'vm_Mgmt'
    return res


def filter_by_vlan(vlan, src_list):
    filtered_rows = list(filter(lambda i: i['vlan'] == vlan, src_list))
    return filtered_rows


def check_psm(src):
    res = False
    if src['vm'] is not None:
        res = re.search('^psm', src['vm'])
    return res


def check_pre(src):
    res = False
    if src['vm'] is not None:
        res = re.search('^pre', src['vm'])
    return res


def get_psm_vip(srv, s_list):
    psm_vip = {}
    for s in s_list:
        if re.search(srv + ".*\(VRRP VIP\)", s['vm']):
            vlan = re.search('\D*', s['vlan']).group(0)
            psm_vip[vlan] = s['ip']
    return psm_vip


def get_sw_list(sheet, dom_num):
    wsid = wb.sheetnames.index('IP-plan')
    net = wb.defined_names.get(sheet.lower(), wsid).attr_text
    ws = wb[str(net[1:net.find('!') - 1])]
    rng = net[net.find('!') + 1:]
    sw_list = []
    for row in ws[rng]:
        if row[0].value == 'OOB_Mgmt':
            sw_list.append({'swname': f'{sheet[:3].upper()}-TMS-1-{dom_num}', 'ip': row[6].value, 'site': 'Site1'})
            sw_list.append({'swname': f'{sheet[:3].upper()}-TMS-2-{dom_num}', 'ip': row[9].value, 'site': 'Site2'})
    return sw_list