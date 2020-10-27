import openpyxl
import yaml
import re

ip_plan = 'project_files/Tele2_IP_plan_v3.02.xlsx'
vars_dir = 'c:/temp/host_vars/'

# Open Excel file in read-only mode
wb = openpyxl.load_workbook(ip_plan, True)

mr = ['SPB', 'MOS', 'NIN', 'EKT', 'NSK']
#mr = ['SPB', 'MOS', 'ROS', 'NIN', 'EKT', 'NSK']
quorum = 1


def get_sheets_list(region):
    ws_list = list(filter(lambda i: re.search('^'+region, i), wb.sheetnames))
    return ws_list


def get_rb_vips(sheet):
    rb_vips = {}
    ws = wb[sheet.upper()]
    for row in ws.iter_rows():
        if re.search("^rb\d\d.*\(VRRP VIP\)", str(row[1].value)) and row[3].value == 'Radius':
            rb_vips[re.search("^rb\d{2}", str(row[1].value)).group(0) + '_vip'] = row[5].value
    return rb_vips


def get_vlans():
    vlans = []
    ws = wb['Dictionary']
    i = 2
    cell = 'D' + str(i)
    while ws[cell].value:
        if 'FlowControl' not in ws[cell].value:
            vlans.append(ws[cell].value)
        i = i + 1
        cell = 'D' + str(i)
    return vlans


def get_prefix_list(sheet):
    vlan_list = get_vlans()
    prefixes = {}
    wsid = wb.sheetnames.index('IP-plan')
    net = wb.defined_names.get(sheet.lower(), wsid).attr_text
    ws = wb[str(net[1:net.find('!') - 1])]
    rng = net[net.find('!') + 1:]
    for row in ws.iter_rows():
        if row[0].value in vlan_list:
            prefixes[row[0].value] = row[3].value
    return prefixes


def get_psm_list(sheet):
    ws = wb[sheet.upper()]
    psm_rows = []
    for row in ws.iter_rows():
        if re.search("^psm[0-9]+", str(row[1].value)):
            psm_rows.append({
                'vm_name': row[1].value,
                'vlan': row[3].value,
                'ip': row[5].value
                })
    return psm_rows


def get_psm_vars(curr_srv, srv_list):
    global quorum
    global prefix_list
    ha = {}
    ha['cluster_id'] = int(re.search('\d{1,2}', curr_srv['vm_name']).group(0))
    ha['quorum'] = quorum
    ha['vip'] = {}
    vars = {}
    vars['cluster'] = {}
    vars['net'] = {}
    for srv in srv_list:
        if srv['vm_name'] == curr_srv['vm_name'] and srv['vlan'] != 'vm_Mgmt':
            vars['net'][re.search("^\D{1,20}", srv['vlan']).group(0).lower() + '_ip'] = srv['ip']
        elif srv['vm_name'] == curr_srv['vm_name'][:-14] + ' (VRRP VIP)':
            ha['vip'][re.search("^\D{1,20}", srv['vlan']).group(0).lower() + '_vip'] = f"{srv['ip']}{prefix_list[srv['vlan']]}"
    vars['cluster']['ha'] = ha
    psm_vars = {curr_srv['vm_name'][:-13]: vars}
    return psm_vars


for region in mr:
    print(f'Collecting data about PSMs in {region}')
    sheet_list = get_sheets_list(region)
    for sheet in sheet_list:
        psm_list = get_psm_list(sheet)
        rb = get_rb_vips(sheet)
        prefix_list = get_prefix_list(sheet)
        for psm in list(filter(lambda i: not re.search('\(VRRP VIP\)', i['vm_name']), psm_list)):
            vars = get_psm_vars(psm, psm_list)
            var_file = f'{vars_dir}{psm["vm_name"][:10]}.yml'
            with open(var_file, 'w', newline='\n') as f:
                f.write(f'# Variables for {psm["vm_name"]}\n#\n')
                f.write('# High availability related vars\n#\n')
                f.write(yaml.dump(vars[psm['vm_name'][:-13]]['cluster']))
            with open(var_file, 'a', newline='\n') as f:
                f.write('#\n' + '# NICs related vars\n#\n')
                f.write(yaml.dump(vars[psm['vm_name'][:-13]]['net']))
            with open(var_file, 'a', newline='\n') as f:
                f.write('#\n' + '# Other vars\n#\n')
                f.write(yaml.dump(rb))
print('Done')